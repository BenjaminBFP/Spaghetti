"""
Carte interactive – rectangles + connexions orthogonales
Drag fluide : l'objet suit le curseur sans téléportation.
"""

import numpy as np
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
from matplotlib.widgets import Button
import json, os, math
import tkinter as tk
from tkinter import colorchooser, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import matplotlib.ticker as mticker

# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTES
# ═══════════════════════════════════════════════════════════════════════════════
LONGUEUR_M   = 100
LARGEUR_M    = 46
TAILLE_CASE  = 0.25
NB_X = int(LONGUEUR_M  / TAILLE_CASE)   # 400
NB_Y = int(LARGEUR_M   / TAILLE_CASE)   # 184

_DIR       = os.path.dirname(os.path.abspath(__file__))
JSON_RECTS = os.path.join(_DIR, "rectangles.json")
JSON_CONN  = os.path.join(_DIR, "connexions.json")
XLSX_FILE  = os.path.join(_DIR, "distances.xlsx")

GAP_ENTRE_TRAITS = 4
BORDER_MARGIN    = 3 
PRODUITS         = ["", "A", "B", "C", "D"]
PRODUIT_COULEURS = {"A": "#4FC3F7", "B": "#E53935", "C": "#43A047", "D": "#FFFFFF", "": "#FF6B6B"}
COULEURS_DEFAUT  = ["#4FC3F7", "#81C784", "#FFB74D", "#E57373", "#BA68C8",
                    "#4DD0E1", "#AED581", "#FFD54F"]
_color_idx = [0]

HANDLE_SIZE = 3
ANCHOR_TOL  = 6   # conservé pour compatibilité, mais les hit-tests utilisent _tol_data()

# Conversion cases ↔ cm (pour l'interface utilisateur)
CASE_TO_CM_X = 40.0  / NB_X   # 0.10 cm/case
CASE_TO_CM_Y = 18.4  / NB_Y   # 0.10 cm/case

# ═══════════════════════════════════════════════════════════════════════════════
#  DONNÉES GLOBALES
# ═══════════════════════════════════════════════════════════════════════════════
rectangles = []
connexions = []
conn_lines = {}
_rects_idx = {}
patches      = {}
texts        = {}
hatch_patches = {}
routes_cache  = {}

state = {
    "selected"          : None,
    "dragging"          : False,
    "drag_offset"       : (0, 0),
    "resizing"          : False,
    "resize_handle"     : None,
    "resize_origin"     : None,
    "handles"           : [],
    "connecting"        : False,
    "conn_first"        : None,
    "conn_selected"     : None,
    "anchor_dragging"   : False,
    "anchor_conn_id"    : None,
    "anchor_end"        : None,
    "anchor_dots"       : [],
    "overlay"           : [],
    # Drag de coude
    "coude_dragging"    : False,
    "coude_conn_id"     : None,
    "coude_idx"         : None,
    "coude_pts_cache"   : {},
    "seg_pts_origin"    : None,
    # Sélection multiple
    "group"             : set(),
    "group_drag_offsets": {},
    "lasso"             : False,
    "lasso_start"       : None,
    "lasso_rect"        : None,
    # Filtre produits
    "produits_visibles" : {"A", "B", "C", "D", ""},
}

# ═══════════════════════════════════════════════════════════════════════════════
#  UTILITAIRES COORDONNÉES
# ═══════════════════════════════════════════════════════════════════════════════
def mat_to_display(x, y, h):
    return x, NB_Y - y - h

def display_to_mat(dx, dy, h):
    return dx, NB_Y - dy - h

def next_color():
    c = COULEURS_DEFAUT[_color_idx[0] % len(COULEURS_DEFAUT)]
    _color_idx[0] += 1
    return c

def rect_by_id(rid):
    return _rects_idx.get(rid)

def _rebuild_rects_idx():
    _rects_idx.clear()
    for r in rectangles:
        _rects_idx[r["id"]] = r

_bbox_cache = {}

def rect_display_bbox(r):
    rid = r["id"]
    if rid not in _bbox_cache:
        dx, dy = mat_to_display(r["x"], r["y"], r["largeur"])
        _bbox_cache[rid] = (dx, dy, dx + r["longueur"], dy + r["largeur"])
    return _bbox_cache[rid]

def _invalidate_bbox(rid):
    _bbox_cache.pop(rid, None)

def side_center(r, side):
    x0, y0, x1, y1 = rect_display_bbox(r)
    cx = (x0 + x1) / 2; cy = (y0 + y1) / 2
    return {"top": (cx, y1), "bottom": (cx, y0), "left": (x0, cy), "right": (x1, cy)}[side]

def closest_side(r, mx, my):
    x0, y0, x1, y1 = rect_display_bbox(r)
    dist = {"top": abs(my - y1), "bottom": abs(my - y0),
            "left": abs(mx - x0), "right": abs(mx - x1)}
    return min(dist, key=dist.get)

def side_axis(side):
    return "h" if side in ("top", "bottom") else "v"

# ═══════════════════════════════════════════════════════════════════════════════
#  TOLÉRANCE DYNAMIQUE PIXEL → DONNÉES
#  Convertit N pixels en unités de données selon la taille courante de la fenêtre.
#  Appel différé (ax existe quand les hit-tests sont appelés).
# ═══════════════════════════════════════════════════════════════════════════════
def _tol_data(pixels=10):
    """
    Retourne une tolérance en unités-données équivalant à `pixels` pixels
    dans l'axe le plus défavorable (le plus petit ratio pixel/unité).
    Fonctionne quelle que soit la taille ou le zoom de la fenêtre.
    """
    try:
        inv = ax.transData.inverted()
        p0 = inv.transform((0.0, 0.0))
        px = inv.transform((float(pixels), 0.0))
        py = inv.transform((0.0, float(pixels)))
        dx = abs(px[0] - p0[0])
        dy = abs(py[1] - p0[1])
        # On prend le max pour être inclusif sur les deux axes
        return max(dx, dy, 1e-3)
    except Exception:
        return pixels  # fallback sûr


# ═══════════════════════════════════════════════════════════════════════════════
#  ANCRAGES
# ═══════════════════════════════════════════════════════════════════════════════
def connexions_sur_cote(rid, side):
    return sorted(
        [c for c in connexions
         if (c["r1"] == rid and c["side1"] == side) or (c["r2"] == rid and c["side2"] == side)],
        key=lambda c: c["id"])

def offset_for_conn(conn_id, rid, side):
    group = connexions_sur_cote(rid, side); n = len(group)
    if n == 0: return 0
    idx = next((i for i, c in enumerate(group) if c["id"] == conn_id), 0)
    r = rect_by_id(rid)
    if r is None: return 0
    side_len = r["longueur"] if side_axis(side) == "h" else r["largeur"]
    gap = GAP_ENTRE_TRAITS
    if gap * (n - 1) > side_len * 0.9:
        gap = (side_len * 0.9) / max(n - 1, 1)
    return -gap * (n - 1) / 2 + idx * gap

def anchor_point(conn_id, rid, side):
    r = rect_by_id(rid)
    if r is None: return (0, 0)
    x0, y0, x1, y1 = rect_display_bbox(r)
    off = offset_for_conn(conn_id, rid, side)
    if side == "top":    cx = (x0 + x1) / 2 + off; return (max(x0, min(x1, cx)), y1)
    if side == "bottom": cx = (x0 + x1) / 2 + off; return (max(x0, min(x1, cx)), y0)
    if side == "left":   cy = (y0 + y1) / 2 + off; return (x0, max(y0, min(y1, cy)))
    if side == "right":  cy = (y0 + y1) / 2 + off; return (x1, max(y0, min(y1, cy)))

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTAGE ORTHOGONAL
# ═══════════════════════════════════════════════════════════════════════════════
def segment_crosses_rect(p1, p2, rx0, ry0, rx1, ry1):
    sx0, sy0 = min(p1[0], p2[0]), min(p1[1], p2[1])
    sx1, sy1 = max(p1[0], p2[0]), max(p1[1], p2[1])
    if sx1 < rx0 or sx0 > rx1 or sy1 < ry0 or sy0 > ry1: return False
    if abs(p1[1] - p2[1]) < 1e-6:
        if ry0 < p1[1] < ry1 and not (sx1 <= rx0 or sx0 >= rx1): return True
    if abs(p1[0] - p2[0]) < 1e-6:
        if rx0 < p1[0] < rx1 and not (sy1 <= ry0 or sy0 >= ry1): return True
    return False
def _clamp_pts(pts):
    if len(pts) <= 2:
        return pts
    mn_x, mx_x = BORDER_MARGIN, NB_X - BORDER_MARGIN
    mn_y, mx_y = BORDER_MARGIN, NB_Y - BORDER_MARGIN
    result = [pts[0]]
    for pt in pts[1:-1]:
        result.append((max(mn_x, min(mx_x, pt[0])),
                       max(mn_y, min(mx_y, pt[1]))))
    result.append(pts[-1])
    return result

def route_orthogonal(p1, side1, p2, side2, r1, r2, conn_id=None):
    if conn_id is not None:
        cache_key = (conn_id,
                     r1["x"], r1["y"], r1["longueur"], r1["largeur"],
                     r2["x"], r2["y"], r2["longueur"], r2["largeur"],
                     round(p1[0], 1), round(p1[1], 1),
                     round(p2[0], 1), round(p2[1], 1))
        if cache_key in routes_cache:
            return routes_cache[cache_key]

    EPS    = 1e-6
    MARGIN = max(GAP_ENTRE_TRAITS * 2, 10)
    PAD    = 8

    exits = {"top": (0, 1), "bottom": (0, -1), "left": (-1, 0), "right": (1, 0)}
    d1x, d1y = exits[side1]; d2x, d2y = exits[side2]
    ox1 = p1[0] + d1x * MARGIN; oy1 = p1[1] + d1y * MARGIN
    ox2 = p2[0] + d2x * MARGIN; oy2 = p2[1] + d2y * MARGIN

    _all_bboxes = [(ro["id"], rect_display_bbox(ro)) for ro in rectangles]
    _r1id = r1["id"]; _r2id = r2["id"]

    def dedup(pts):
        out = [pts[0]]
        for pt in pts[1:]:
            if abs(pt[0] - out[-1][0]) > EPS or abs(pt[1] - out[-1][1]) > EPS:
                out.append(pt)
        return out

    def clean_colinear(pts):
        if len(pts) <= 2: return pts
        out = [pts[0]]
        for i in range(1, len(pts) - 1):
            ax_, ay_ = out[-1]; bx, by = pts[i]; cx_, cy_ = pts[i + 1]
            if not ((abs(ax_ - bx) < EPS and abs(bx - cx_) < EPS) or
                    (abs(ay_ - by) < EPS and abs(by - cy_) < EPS)):
                out.append(pts[i])
        out.append(pts[-1])
        return out

    def ortho_path(raw_waypoints):
        result = [raw_waypoints[0]]
        for i in range(len(raw_waypoints) - 1):
            a = result[-1]; b = raw_waypoints[i + 1]
            dx = abs(b[0] - a[0]); dy = abs(b[1] - a[1])
            if dx < EPS or dy < EPS:
                result.append(b)
            else:
                c1 = (b[0], a[1]); c2 = (a[0], b[1])
                def seg_hits(pa, pb):
                    for (_, bbox) in _all_bboxes:
                        if segment_crosses_rect(pa, pb, *bbox): return 1
                    return 0
                if seg_hits(a, c1) + seg_hits(c1, b) <= seg_hits(a, c2) + seg_hits(c2, b):
                    result.append(c1)
                else:
                    result.append(c2)
                result.append(b)
        return clean_colinear(dedup(result))

    def path_length(pts):
        return sum(math.sqrt((pts[i + 1][0] - pts[i][0]) ** 2 +
                             (pts[i + 1][1] - pts[i][1]) ** 2)
                   for i in range(len(pts) - 1))

    def path_collisions(pts):
        count = 0; n = len(pts) - 1
        for i in range(n):
            a = pts[i]; b = pts[i + 1]
            for (rid_ro, bbox) in _all_bboxes:
                if rid_ro == _r1id and i == 0: continue
                if rid_ro == _r2id and i == n - 1: continue
                if segment_crosses_rect(a, b, *bbox):
                    count += 1; break
        return count

    candidates = []
    candidates.append(ortho_path([p1, (ox1, oy1), (ox2, oy2), p2]))
    for mid in [(ox2, oy1), (ox1, oy2)]:
        candidates.append(ortho_path([p1, (ox1, oy1), mid, (ox2, oy2), p2]))
    zone_x0 = min(ox1, ox2) - PAD * 3; zone_y0 = min(oy1, oy2) - PAD * 3
    zone_x1 = max(ox1, ox2) + PAD * 3; zone_y1 = max(oy1, oy2) + PAD * 3
    for ro in rectangles:
        rx0, ry0, rx1, ry1 = rect_display_bbox(ro)
        if rx1 < zone_x0 or rx0 > zone_x1 or ry1 < zone_y0 or ry0 > zone_y1: continue
        for wy in (ry0 - PAD, ry1 + PAD):
            candidates.append(ortho_path([p1, (ox1, oy1), (ox1, wy), (ox2, wy), (ox2, oy2), p2]))
        for wx in (rx0 - PAD, rx1 + PAD):
            candidates.append(ortho_path([p1, (ox1, oy1), (wx, oy1), (wx, oy2), (ox2, oy2), p2]))

    best = None; best_score = (999999, 999999.0)
    for pts in candidates:
        score = (path_collisions(pts), path_length(pts))
        if score < best_score:
            best_score = score; best = pts

    result = _clamp_pts(best if best is not None else candidates[0])
    if conn_id is not None:
        routes_cache[cache_key] = result
    return result

def pts_depuis_override(conn, p1, p2):
    inner = conn["waypoints_override"][1:-1]
    if not inner: return None
    EPS2 = 1e-6
    exits = {"top": (0, 1), "bottom": (0, -1), "left": (-1, 0), "right": (1, 0)}
    d1x, d1y = exits[conn["side1"]]; d2x, d2y = exits[conn["side2"]]
    MARGIN = max(GAP_ENTRE_TRAITS * 2, 10)
    ox1 = p1[0] + d1x * MARGIN; oy1 = p1[1] + d1y * MARGIN
    ox2 = p2[0] + d2x * MARGIN; oy2 = p2[1] + d2y * MARGIN

    def ortho2(a, b):
        if abs(a[0] - b[0]) < EPS2 or abs(a[1] - b[1]) < EPS2: return [b]
        return [(b[0], a[1]), b]

    seg = [p1, (ox1, oy1)]
    seg += ortho2((ox1, oy1), inner[0])
    for i in range(len(inner) - 1): seg += ortho2(inner[i], inner[i + 1])
    seg += ortho2(inner[-1], (ox2, oy2))
    seg += [(ox2, oy2), p2]

    out = [seg[0]]
    for pt in seg[1:]:
        if abs(pt[0] - out[-1][0]) > EPS2 or abs(pt[1] - out[-1][1]) > EPS2:
            out.append(pt)
    clean = [out[0]]
    for i in range(1, len(out) - 1):
        ax_, ay_ = clean[-1]; bx, by = out[i]; cx_, cy_ = out[i + 1]
        if not ((abs(ax_ - bx) < EPS2 and abs(bx - cx_) < EPS2) or
                (abs(ay_ - by) < EPS2 and abs(by - cy_) < EPS2)):
            clean.append(out[i])
    clean.append(out[-1])
    return clean

def distance_connexion(conn):
    r1 = rect_by_id(conn["r1"]); r2 = rect_by_id(conn["r2"])
    p1 = anchor_point(conn["id"], conn["r1"], conn["side1"])
    p2 = anchor_point(conn["id"], conn["r2"], conn["side2"])
    if "waypoints_override" in conn:
        pts = pts_depuis_override(conn, p1, p2)
        if pts is None:
            pts = route_orthogonal(p1, conn["side1"], p2, conn["side2"], r1, r2, conn_id=conn["id"])
    elif "pts_cache" in conn:
        pts = conn["pts_cache"]
    else:
        pts = route_orthogonal(p1, conn["side1"], p2, conn["side2"], r1, r2, conn_id=conn["id"])
        conn["pts_cache"] = pts
    total = sum(math.sqrt((pts[i + 1][0] - pts[i][0]) ** 2 +
                          (pts[i + 1][1] - pts[i][1]) ** 2)
                for i in range(len(pts) - 1))
    return total * TAILLE_CASE

# ═══════════════════════════════════════════════════════════════════════════════
#  SAUVEGARDE / CHARGEMENT
# ═══════════════════════════════════════════════════════════════════════════════
def sauvegarder(avec_xlsx=True):
    # ── Rectangles : on sauvegarde l'id pour éviter le décalage au rechargement ──
    export = [{"id": r["id"], "nom": r["nom"],
               "longueur": r["longueur"], "largeur": r["largeur"],
               "position": {"x": r["x"], "y": r["y"]}, "couleur": r["couleur"],
               "orientation_texte": r["orientation"],
               "lock_x": r.get("lock_x", False), "lock_y": r.get("lock_y", False),
               "rayures": r.get("rayures", False),
               "couleur_rayures": r.get("couleur_rayures", "#FFFFFF"),
               "couleur_texte": r.get("couleur_texte", "white")}
              for r in rectangles]
    with open(JSON_RECTS, "w") as f: json.dump(export, f, indent=4)

    conn_export = []
    for c in connexions:
        # On exclut pts_cache (objet Python) et on le sérialise séparément
        # sous la clé "route_sauvegardee" (liste de listes, compatible JSON).
        # waypoints_override est conservé (formes manuelles persistantes).
        ce = {k: v for k, v in c.items() if k != "pts_cache"}
        if "pts_cache" in c and c["pts_cache"]:
            ce["route_sauvegardee"] = [list(p) for p in c["pts_cache"]]
        conn_export.append(ce)
    with open(JSON_CONN, "w") as f: json.dump(conn_export, f, indent=4)

    if avec_xlsx:
        exporter_xlsx()

def charger():
    global rectangles, connexions
    if os.path.exists(JSON_RECTS):
        with open(JSON_RECTS) as f: data = json.load(f)
        rectangles = []
        for i, d in enumerate(data):
            rectangles.append({
                # Utilise l'id sauvegardé → les connexions pointent toujours
                # vers les bons blocs même après suppression/ajout.
                "id": d.get("id", i),
                "nom": d["nom"],
                "longueur": d["longueur"], "largeur": d["largeur"],
                "x": d["position"]["x"], "y": d["position"]["y"],
                "couleur": d["couleur"],
                "orientation": d.get("orientation_texte", "horizontal"),
                "lock_x": d.get("lock_x", False), "lock_y": d.get("lock_y", False),
                "rayures": d.get("rayures", False),
                "couleur_rayures": d.get("couleur_rayures", "#FFFFFF"),
                "couleur_texte": d.get("couleur_texte", "white"),
            })
    if os.path.exists(JSON_CONN):
        with open(JSON_CONN) as f: connexions = json.load(f)
        for c in connexions:
            # Restaurer le chemin calculé sauvegardé → pas de recalcul à l'ouverture.
            # Les coordonnées sont stables car les positions des blocs sont aussi
            # sauvegardées et la transformation mat_to_display est déterministe.
            if "route_sauvegardee" in c:
                c["pts_cache"] = [tuple(p) for p in c["route_sauvegardee"]]
                del c["route_sauvegardee"]
            else:
                c.pop("pts_cache", None)
            # waypoints_override : NE PAS effacer — il est maintenant sauvegardé
            # et représente la forme manuelle choisie par l'utilisateur.
            # On ne le supprime que s'il vient d'un ancien fichier sans "pts_cache"
            # associé (format pré-correction), détectable à l'absence de route_sauvegardee.

def exporter_xlsx():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Distances"
    hdr_fill  = PatternFill("solid", fgColor="1565C0")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    cell_font  = Font(name="Arial", size=10)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers    = ["Point A", "Point B", "Distance (m)", "Nombre", "Produit", "Conversion", "Distance × Nombre"]
    col_widths = [20, 20, 16, 12, 14, 14, 20]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    alt_fill = PatternFill("solid", fgColor="E3F2FD")
    for i, conn in enumerate(connexions, 2):
        r1 = rect_by_id(conn["r1"]); r2 = rect_by_id(conn["r2"])
        nom_a = r1["nom"] if r1 else str(conn["r1"])
        nom_b = r2["nom"] if r2 else str(conn["r2"])
        dist = round(distance_connexion(conn), 3)
        nombre = conn.get("nombre", 1); produit = conn.get("produit", "")
        row_fill = alt_fill if i % 2 == 0 else None
        for ci, v in enumerate([nom_a, nom_b, dist, nombre, produit, "", ""], 1):
            cell = ws.cell(row=i, column=ci, value=v)
            cell.font = cell_font; cell.alignment = cell_align; cell.border = border
            if row_fill: cell.fill = row_fill
        cc = openpyxl.utils.get_column_letter(3); cd = openpyxl.utils.get_column_letter(4)
        ws.cell(row=i, column=7).value = f"={cc}{i}*{cd}{i}"
        ws.cell(row=i, column=7).font = Font(name="Arial", size=10, color="000000")
        ws.cell(row=i, column=7).alignment = cell_align
        ws.cell(row=i, column=7).border = border
    wb.save(XLSX_FILE)

# ═══════════════════════════════════════════════════════════════════════════════
#  DESSIN
# ═══════════════════════════════════════════════════════════════════════════════
def clear_overlay():
    for artist in state["overlay"]:
        try: artist.remove()
        except: pass
    state["overlay"].clear()

def draw_all():
    _rebuild_rects_idx()
    _bbox_cache.clear()
    clear_overlay()
    cids_visibles = {c["id"] for c in connexions
                     if c.get("produit", "") in state["produits_visibles"]}
    for key in list(routes_cache.keys()):
        if key[0] in cids_visibles:
            del routes_cache[key]
    for p in patches.values(): p.remove()
    for hp in hatch_patches.values(): hp.remove()
    for t in texts.values(): t.remove()
    patches.clear(); hatch_patches.clear(); texts.clear()
    clear_handles(); clear_conn_lines()
    for r in rectangles: draw_rect(r)
    for c in connexions: draw_connexion(c)
    if state["selected"] is not None: draw_handles(state["selected"])
    fig.canvas.draw_idle()

def draw_rect(r):
    rid = r["id"]
    dx, dy = mat_to_display(r["x"], r["y"], r["largeur"])
    sel      = (state["selected"] == rid)
    in_group = (rid in state["group"]) and not sel
    ec = "yellow" if sel else ("#00FFFF" if in_group else "#cccccc")
    lw = 2 if (sel or in_group) else 1
    couleur = r["couleur"]
    if isinstance(couleur, str) and couleur.startswith("#") and len(couleur) not in (4, 7):
        couleur = "#4FC3F7"
    p = Rectangle((dx, dy), r["longueur"], r["largeur"],
                  linewidth=lw, edgecolor=ec, facecolor=couleur, alpha=0.65, picker=True)
    p.rid = rid; ax.add_patch(p); patches[rid] = p
    if r.get("rayures", False):
        hatch_color = r.get("couleur_rayures", "#FFFFFF")
        ph = Rectangle((dx, dy), r["longueur"], r["largeur"],
                        linewidth=0.8, edgecolor=hatch_color, facecolor="none",
                        hatch="///", alpha=0.9, zorder=2)
        ax.add_patch(ph); hatch_patches[rid] = ph
    rot = 0 if r["orientation"] == "horizontal" else 90
    lock_lx = r.get("lock_x", False); lock_ly = r.get("lock_y", False)
    lock_str = ""
    if lock_lx and lock_ly: lock_str = " [XY]"
    elif lock_lx: lock_str = " [X]"
    elif lock_ly: lock_str = " [Y]"
    full_text = r["nom"] + lock_str
    side_len  = r["longueur"] if rot == 0 else r["largeur"]

    def wrap_text(text, max_width_cases):
        CHAR_W = 2.0
        words = text.split(); lines_out = []; line = ""
        for w in words:
            candidate = (line + (" " if line else "") + w)
            if len(candidate) * CHAR_W <= max_width_cases:
                line = candidate
            else:
                if line: lines_out.append(line)
                line = w
        if line: lines_out.append(line)
        return "\n".join(lines_out)

    display_text = wrap_text(full_text, side_len)
    txt_color = r.get("couleur_texte", "white")
    t = ax.text(dx + r["longueur"] / 2, dy + r["largeur"] / 2, display_text,
                ha="center", va="center", fontsize=7, rotation=rot, clip_on=True,
                color=txt_color, fontweight="bold", linespacing=1.3)
    texts[rid] = t

def draw_connexion(conn):
    cid = conn["id"]
    r1  = rect_by_id(conn["r1"]); r2 = rect_by_id(conn["r2"])
    if r1 is None or r2 is None: return
    if conn.get("produit", "") not in state["produits_visibles"]: return
    p1 = anchor_point(cid, conn["r1"], conn["side1"])
    p2 = anchor_point(cid, conn["r2"], conn["side2"])
    if "waypoints_override" in conn:
        pts = pts_depuis_override(conn, p1, p2)
        if pts is None:
            pts = route_orthogonal(p1, conn["side1"], p2, conn["side2"], r1, r2, conn_id=cid)
    elif "pts_cache" in conn:
        pts = conn["pts_cache"]
    else:
        pts = route_orthogonal(p1, conn["side1"], p2, conn["side2"], r1, r2, conn_id=cid)
        conn["pts_cache"] = pts
    sel        = (state["conn_selected"] == cid)
    prod       = conn.get("produit", "")
    base_color = PRODUIT_COULEURS.get(prod, "#FF6B6B")
    color = "#FFD700" if sel else base_color; lw = 2.5 if sel else 1.8
    lines = []
    for i in range(len(pts) - 1):
        ln, = ax.plot([pts[i][0], pts[i + 1][0]], [pts[i][1], pts[i + 1][1]],
                      color=color, linewidth=lw, solid_capstyle="round", zorder=5)
        ln.conn_id = cid; lines.append(ln)
    mid  = len(pts) // 2
    mx_l = (pts[mid - 1][0] + pts[mid][0]) / 2
    my_l = (pts[mid - 1][1] + pts[mid][1]) / 2
    nb = conn.get("nombre", 1)
    if state["coude_dragging"] and state["coude_conn_id"] == cid:
        label_txt = f"...m  x{nb}"
    else:
        label_txt = f"{distance_connexion(conn):.1f}m  x{nb}"
    lt = ax.text(mx_l, my_l, label_txt, ha="center", va="bottom", fontsize=6,
                 color="#FFD700", fontweight="bold",
                 bbox=dict(boxstyle="round,pad=0.2", fc="#1a1a2e", ec="none", alpha=0.7),
                 zorder=4, picker=False)
    lines.append(lt)
    dot_color = "#FFFFFF" if sel else base_color; dot_size = 4 if sel else 2.5
    for pt in (p1, p2):
        dot, = ax.plot(pt[0], pt[1], "s", color=dot_color, markersize=dot_size, zorder=8)
        dot.conn_id = cid; lines.append(dot)
    if sel:
        state["coude_pts_cache"][cid] = pts
        seg_range = range(0, len(pts) - 1) if len(pts) <= 3 else range(1, len(pts) - 2)
        for i in seg_range:
            mx_seg = (pts[i][0] + pts[i + 1][0]) / 2
            my_seg = (pts[i][1] + pts[i + 1][1]) / 2
            cd, = ax.plot(mx_seg, my_seg, "D", color="#FF9800", markersize=7, zorder=12,
                          markeredgecolor="white", markeredgewidth=0.8)
            cd.conn_id = cid; cd.seg_idx = i; lines.append(cd)
    else:
        state["coude_pts_cache"].pop(cid, None)
    conn_lines[cid] = lines

def clear_conn_lines():
    for items in conn_lines.values():
        for item in items: item.remove()
    conn_lines.clear()

def draw_handles(rid):
    r = rect_by_id(rid)
    if r is None: return
    dx, dy = mat_to_display(r["x"], r["y"], r["largeur"]); s = HANDLE_SIZE
    for hx, hy, hname in [
        (dx + r["longueur"] - s / 2, dy + r["largeur"] / 2 - s / 2, "right"),
        (dx + r["longueur"] / 2 - s / 2, dy - s / 2, "bottom"),
        (dx + r["longueur"] - s / 2, dy - s / 2, "corner"),
    ]:
        hp = Rectangle((hx, hy), s, s, linewidth=1,
                        edgecolor="white", facecolor="yellow", alpha=0.9, zorder=10)
        hp.handle_name = hname; ax.add_patch(hp); state["handles"].append(hp)

def clear_handles():
    for hp in state["handles"]: hp.remove()
    state["handles"].clear()

# ═══════════════════════════════════════════════════════════════════════════════
#  REDRAW PARTIEL (drag fluide)
# ═══════════════════════════════════════════════════════════════════════════════
def _partial_redraw_moved(moved_rids):
    _rebuild_rects_idx()
    for rid_m in moved_rids:
        rm = rect_by_id(rid_m)
        if rm is None: continue
        _invalidate_bbox(rid_m)
        dx_p, dy_p = mat_to_display(rm["x"], rm["y"], rm["largeur"])
        if rid_m in patches:
            patches[rid_m].set_xy((dx_p, dy_p))
        if rid_m in hatch_patches:
            hatch_patches[rid_m].set_xy((dx_p, dy_p))
        if rid_m in texts:
            texts[rid_m].set_position((dx_p + rm["longueur"] / 2,
                                       dy_p + rm["largeur"] / 2))
    touched_cids = [c["id"] for c in connexions
                    if (c["r1"] in moved_rids or c["r2"] in moved_rids)
                    and c.get("produit", "") in state["produits_visibles"]]
    for cid in touched_cids:
        if cid in conn_lines:
            for item in conn_lines[cid]:
                try: item.remove()
                except: pass
            del conn_lines[cid]
        conn = next((c for c in connexions if c["id"] == cid), None)
        if conn: draw_connexion(conn)
    fig.canvas.draw_idle()

# ═══════════════════════════════════════════════════════════════════════════════
#  FENÊTRES TKINTER
# ═══════════════════════════════════════════════════════════════════════════════
def _safe_preview(label, var):
    try: label.configure(bg=var.get())
    except: pass

def editer_connexion(cid):
    conn = next((c for c in connexions if c["id"] == cid), None)
    if conn is None: return
    r1 = rect_by_id(conn["r1"]); r2 = rect_by_id(conn["r2"])
    nom_a = r1["nom"] if r1 else "?"; nom_b = r2["nom"] if r2 else "?"
    BG = "#1e1e2e"; BG2 = "#2a2a3e"; FG = "#e0e0e0"; ACCENT = "#FF6B6B"
    FONT = ("Segoe UI", 10); FONT_B = ("Segoe UI", 10, "bold")
    win = tk.Toplevel(); win.title("Editer connexion"); win.resizable(False, False)
    win.configure(bg=BG); win.grab_set()
    tk.Label(win, text=f"Connexion : {nom_a}  →  {nom_b}", bg=BG, fg=ACCENT,
             font=("Segoe UI", 11, "bold"), pady=10).grid(row=0, column=0, columnspan=2,
                                                           sticky="ew", padx=16)
    tk.Frame(win, bg=ACCENT, height=1).grid(row=1, column=0, columnspan=2,
                                             sticky="ew", padx=16, pady=(0, 10))
    tk.Label(win, text="Nombre (entier) :", bg=BG, fg=FG, font=FONT,
             anchor="w", width=20).grid(row=2, column=0, sticky="w", padx=(16, 4), pady=8)
    nb_var = tk.StringVar(value=str(conn.get("nombre", 1)))
    e = tk.Entry(win, textvariable=nb_var, bg=BG2, fg=FG, insertbackground=FG,
                 relief="flat", font=FONT, width=10, highlightthickness=1,
                 highlightbackground="#555", highlightcolor=ACCENT)
    e.grid(row=2, column=1, padx=(0, 16), pady=8, sticky="w")
    tk.Label(win, text="Produit :", bg=BG, fg=FG, font=FONT,
             anchor="w", width=20).grid(row=3, column=0, sticky="w", padx=(16, 4), pady=8)
    prod_var = tk.StringVar(value=conn.get("produit", ""))
    prod_frame = tk.Frame(win, bg=BG); prod_frame.grid(row=3, column=1, sticky="w", pady=8)
    for val, label, col in [("", "Aucun", "#888888"), ("A", "A", "#4FC3F7"),
                             ("B", "B", "#E53935"), ("C", "C", "#43A047"), ("D", "D", "#FFFFFF")]:
        tk.Radiobutton(prod_frame, text=label, variable=prod_var, value=val,
                       bg=BG, fg=col, selectcolor=BG2, activebackground=BG,
                       activeforeground=col, font=FONT_B, indicatoron=1).pack(side="left", padx=4)
    tk.Frame(win, bg="#444", height=1).grid(row=4, column=0, columnspan=2,
                                             sticky="ew", padx=16, pady=(6, 4))
    def valider():
        try:
            n = int(nb_var.get())
            if n < 1: raise ValueError
        except ValueError:
            messagebox.showerror("Erreur", "Le nombre doit etre un entier >= 1.", parent=win); return
        conn["nombre"] = n; conn["produit"] = prod_var.get()
        win.destroy(); draw_all(); sauvegarder()
    bf = tk.Frame(win, bg=BG); bf.grid(row=5, column=0, columnspan=2, pady=12)
    bs = dict(relief="flat", font=FONT_B, padx=16, pady=6, cursor="hand2", bd=0)
    tk.Button(bf, text="✅ Valider", command=valider, bg="#2e7d32", fg="white",
              activebackground="#388e3c", activeforeground="white", **bs).pack(side="left", padx=8)
    tk.Button(bf, text="✖ Annuler", command=win.destroy, bg="#b71c1c", fg="white",
              activebackground="#c62828", activeforeground="white", **bs).pack(side="left", padx=8)
    win.bind("<Return>", lambda ev: valider()); win.bind("<Escape>", lambda ev: win.destroy())
    win.update_idletasks()
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    win.geometry(f"+{(sw - win.winfo_width()) // 2}+{(sh - win.winfo_height()) // 2}")
    e.focus_set(); e.select_range(0, tk.END)

def ouvrir_fenetre_rect(titre, defaults, callback, groupe=False):
    BG = "#1e1e2e"; BG2 = "#2a2a3e"; FG = "#e0e0e0"; ACCENT = "#4FC3F7"
    FONT = ("Segoe UI", 10); FONT_B = ("Segoe UI", 10, "bold")
    win = tk.Toplevel(); win.title(titre); win.resizable(False, False)
    win.configure(bg=BG); win.grab_set()
    tk.Label(win, text=titre, bg=BG, fg=ACCENT,
             font=("Segoe UI", 12, "bold"), pady=10).grid(row=0, column=0, columnspan=3,
                                                           sticky="ew", padx=16)
    tk.Frame(win, bg=ACCENT, height=1).grid(row=1, column=0, columnspan=3,
                                             sticky="ew", padx=16, pady=(0, 8))
    entries = {}

    def add_field(row, label, key, value, tip=""):
        tk.Label(win, text=label, bg=BG, fg=FG, font=FONT, anchor="w",
                 width=24).grid(row=row, column=0, sticky="w", padx=(16, 4), pady=5)
        e = tk.Entry(win, bg=BG2, fg=FG, insertbackground=FG, relief="flat",
                     font=FONT, width=16, highlightthickness=1,
                     highlightbackground="#555", highlightcolor=ACCENT)
        e.insert(0, str(value)); e.grid(row=row, column=1, padx=(0, 16), pady=5, sticky="w")
        entries[key] = e
        if tip:
            tk.Label(win, text=tip, bg=BG, fg="#888",
                     font=("Segoe UI", 8)).grid(row=row, column=2, sticky="w", padx=(0, 16))

    add_field(2, "Nom du rectangle", "nom", defaults["nom"])
    add_field(3, "Longueur (cm)", "longueur", defaults["longueur"],
              f"max {NB_X * CASE_TO_CM_X:.1f} cm")
    add_field(4, "Largeur  (cm)", "largeur", defaults["largeur"],
              f"max {NB_Y * CASE_TO_CM_Y:.1f} cm")
    add_field(5, "X coin sup-gauche (cm)", "x", defaults["x"],
              f"0 – {NB_X * CASE_TO_CM_X:.1f}")
    add_field(6, "Y coin sup-gauche (cm)", "y", defaults["y"],
              f"0 – {NB_Y * CASE_TO_CM_Y:.1f}")

    tk.Label(win, text="Couleur", bg=BG, fg=FG, font=FONT, anchor="w",
             width=24).grid(row=7, column=0, sticky="w", padx=(16, 4), pady=5)
    color_var = tk.StringVar(value=defaults["couleur"])
    cf = tk.Frame(win, bg=BG); cf.grid(row=7, column=1, sticky="w", pady=5)
    preview = tk.Label(cf, bg=defaults["couleur"], width=3, height=1, relief="solid", bd=1)
    preview.pack(side="left", padx=(0, 5))
    tk.Entry(cf, textvariable=color_var, bg=BG2, fg=FG, insertbackground=FG,
             relief="flat", font=FONT, width=10, highlightthickness=1,
             highlightbackground="#555", highlightcolor=ACCENT).pack(side="left")
    def pick():
        try: res = colorchooser.askcolor(color=color_var.get(), parent=win, title="Couleur")
        except: res = (None, None)
        if res[1]: color_var.set(res[1])
    color_var.trace_add("write", lambda *_: _safe_preview(preview, color_var))
    tk.Button(cf, text="🎨", command=pick, bg=BG2, fg=FG, relief="flat",
              font=FONT, bd=0, cursor="hand2").pack(side="left", padx=(5, 0))

    tk.Label(win, text="Rayures diagonales", bg=BG, fg=FG, font=FONT, anchor="w",
             width=24).grid(row=8, column=0, sticky="w", padx=(16, 4), pady=5)
    rayures_var = tk.BooleanVar(value=defaults.get("rayures", False))
    rf = tk.Frame(win, bg=BG); rf.grid(row=8, column=1, sticky="w", pady=5)
    tk.Checkbutton(rf, text="Activer", variable=rayures_var, bg=BG, fg=FG,
                   selectcolor=BG2, activebackground=BG, activeforeground=ACCENT,
                   font=FONT).pack(side="left", padx=(0, 10))
    couleur_rayures_var = tk.StringVar(value=defaults.get("couleur_rayures", "#FFFFFF"))
    preview_r = tk.Label(rf, bg=defaults.get("couleur_rayures", "#FFFFFF"),
                         width=3, height=1, relief="solid", bd=1)
    preview_r.pack(side="left", padx=(0, 5))
    tk.Entry(rf, textvariable=couleur_rayures_var, bg=BG2, fg=FG, insertbackground=FG,
             relief="flat", font=FONT, width=9, highlightthickness=1,
             highlightbackground="#555", highlightcolor=ACCENT).pack(side="left")
    def pick_rayures():
        try: res = colorchooser.askcolor(color=couleur_rayures_var.get(), parent=win, title="Couleur rayures")
        except: res = (None, None)
        if res[1]: couleur_rayures_var.set(res[1])
    couleur_rayures_var.trace_add("write", lambda *_: _safe_preview(preview_r, couleur_rayures_var))
    tk.Button(rf, text="🎨", command=pick_rayures, bg=BG2, fg=FG, relief="flat",
              font=FONT, bd=0, cursor="hand2").pack(side="left", padx=(5, 0))

    tk.Label(win, text="Couleur du texte", bg=BG, fg=FG, font=FONT, anchor="w",
             width=24).grid(row=9, column=0, sticky="w", padx=(16, 4), pady=5)
    couleur_texte_var = tk.StringVar(value=defaults.get("couleur_texte", "white"))
    ctf = tk.Frame(win, bg=BG); ctf.grid(row=9, column=1, sticky="w", pady=5)
    tk.Radiobutton(ctf, text="Blanc", variable=couleur_texte_var, value="white",
                   bg=BG, fg="#FFFFFF", selectcolor=BG2, activebackground=BG,
                   activeforeground="#FFFFFF", font=FONT).pack(side="left", padx=(0, 10))
    tk.Radiobutton(ctf, text="Noir", variable=couleur_texte_var, value="black",
                   bg=BG, fg="#AAAAAA", selectcolor=BG2, activebackground=BG,
                   activeforeground="#AAAAAA", font=FONT).pack(side="left")

    tk.Label(win, text="Orientation texte", bg=BG, fg=FG, font=FONT, anchor="w",
             width=24).grid(row=10, column=0, sticky="w", padx=(16, 4), pady=5)
    orient_var = tk.StringVar(value=defaults["orientation"])
    of = tk.Frame(win, bg=BG); of.grid(row=10, column=1, sticky="w", pady=5)
    for v, l in [("horizontal", "⟷ Horizontal"), ("vertical", "↕ Vertical")]:
        tk.Radiobutton(of, text=l, variable=orient_var, value=v, bg=BG, fg=FG,
                       selectcolor=BG2, activebackground=BG, activeforeground=ACCENT,
                       font=FONT).pack(side="left", padx=6)

    lock_frame = tk.Frame(win, bg=BG)
    lock_frame.grid(row=11, column=0, columnspan=3, sticky="w", padx=16, pady=(4, 2))
    tk.Label(lock_frame, text="Verrouiller position :", bg=BG, fg=FG,
             font=FONT).pack(side="left", padx=(0, 10))
    lock_x_var = tk.BooleanVar(value=defaults.get("lock_x", False))
    lock_y_var = tk.BooleanVar(value=defaults.get("lock_y", False))
    tk.Checkbutton(lock_frame, text="🔒 X  (horizontal)", variable=lock_x_var,
                   bg=BG, fg=FG, selectcolor=BG2, activebackground=BG,
                   activeforeground=ACCENT, font=FONT).pack(side="left", padx=(0, 12))
    tk.Checkbutton(lock_frame, text="🔒 Y  (vertical)", variable=lock_y_var,
                   bg=BG, fg=FG, selectcolor=BG2, activebackground=BG,
                   activeforeground=ACCENT, font=FONT).pack(side="left")

    tk.Frame(win, bg="#444", height=1).grid(row=12, column=0, columnspan=3,
                                             sticky="ew", padx=16, pady=(8, 4))

    def valider():
        def parse_or_empty(s):
            s = s.strip()
            if s == "": return ""
            return float(s)
        try:
            lon_raw = parse_or_empty(entries["longueur"].get())
            lar_raw = parse_or_empty(entries["largeur"].get())
            x_raw   = parse_or_empty(entries["x"].get())
            y_raw   = parse_or_empty(entries["y"].get())
            longueur = "" if lon_raw == "" else max(1, round(lon_raw / CASE_TO_CM_X))
            largeur  = "" if lar_raw == "" else max(1, round(lar_raw / CASE_TO_CM_Y))
            x        = "" if x_raw == ""   else max(0, round(x_raw / CASE_TO_CM_X))
            if y_raw == "":
                y_mat = ""
            else:
                larg_for_y = largeur if largeur != "" else 0
                y_affich = round(y_raw / CASE_TO_CM_Y)
                y_mat = NB_Y - y_affich - larg_for_y
            nom = entries["nom"].get().strip()
            if not groupe and not nom: nom = "Rect"
            result = {"nom": nom, "longueur": longueur, "largeur": largeur,
                      "x": x, "y": y_mat, "couleur": color_var.get(),
                      "orientation": orient_var.get(),
                      "lock_x": lock_x_var.get(), "lock_y": lock_y_var.get(),
                      "rayures": rayures_var.get(),
                      "couleur_rayures": couleur_rayures_var.get(),
                      "couleur_texte": couleur_texte_var.get()}
        except ValueError:
            messagebox.showerror("Erreur", "Longueur/Largeur/X/Y doivent être des nombres.", parent=win)
            return
        win.destroy(); callback(result)

    bf = tk.Frame(win, bg=BG); bf.grid(row=13, column=0, columnspan=3, pady=12)
    bs = dict(relief="flat", font=FONT_B, padx=18, pady=7, cursor="hand2", bd=0)
    tk.Button(bf, text="✅ Valider", command=valider, bg="#2e7d32", fg="white",
              activebackground="#388e3c", activeforeground="white", **bs).pack(side="left", padx=8)
    tk.Button(bf, text="✖ Annuler", command=win.destroy, bg="#b71c1c", fg="white",
              activebackground="#c62828", activeforeground="white", **bs).pack(side="left", padx=8)
    win.bind("<Return>", lambda e: valider()); win.bind("<Escape>", lambda e: win.destroy())
    win.update_idletasks()
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    win.geometry(f"+{(sw - win.winfo_width()) // 2}+{(sh - win.winfo_height()) // 2}")
    entries["nom"].focus_set(); entries["nom"].select_range(0, tk.END)

def ajouter_rectangle_interactif(*args):
    defaults = {"nom": "Nouveau rect",
                "longueur": round(40 * CASE_TO_CM_X, 2),
                "largeur":  round(20 * CASE_TO_CM_Y, 2),
                "x": 0.0, "y": 0.0, "couleur": next_color(),
                "orientation": "horizontal",
                "lock_x": False, "lock_y": False,
                "rayures": False, "couleur_rayures": "#FFFFFF", "couleur_texte": "white"}
    def ok(r):
        rid = max((rec["id"] for rec in rectangles), default=-1) + 1
        rectangles.append({"id": rid, **r}); draw_all(); sauvegarder()
    fig.canvas.get_tk_widget().winfo_toplevel().after(
        10, lambda: ouvrir_fenetre_rect("➕ Ajouter un rectangle", defaults, ok))

def editer_groupe():
    rids = list(state["group"])
    if not rids: return
    rects = [rect_by_id(rid) for rid in rids if rect_by_id(rid)]
    if not rects: return

    def shared(key, transform=None):
        vals = [r.get(key) for r in rects]
        if transform: vals = [transform(v) for v in vals]
        return vals[0] if len(set(str(v) for v in vals)) == 1 else ""

    def shared_bool(key):
        vals = [r.get(key, False) for r in rects]
        return vals[0] if len(set(vals)) == 1 else False

    defaults = {
        "nom"            : shared("nom"),
        "longueur"       : shared("longueur", lambda v: round(v * CASE_TO_CM_X, 3)),
        "largeur"        : shared("largeur",  lambda v: round(v * CASE_TO_CM_Y, 3)),
        "x"              : shared("x",        lambda v: round(v * CASE_TO_CM_X, 3)),
        "couleur"        : shared("couleur"),
        "orientation"    : shared("orientation"),
        "lock_x"         : shared_bool("lock_x"),
        "lock_y"         : shared_bool("lock_y"),
        "rayures"        : shared_bool("rayures"),
        "couleur_rayures": shared("couleur_rayures"),
        "couleur_texte"  : shared("couleur_texte"),
    }
    y_vals = set(r["y"] for r in rects)
    if len(y_vals) == 1:
        r0 = rects[0]
        defaults["y"] = round((NB_Y - r0["y"] - r0["largeur"]) * CASE_TO_CM_Y, 3)
    else:
        defaults["y"] = ""

    def ok(result):
        raw = {k: result.get(k) for k in
               ["nom", "longueur", "largeur", "x", "y", "couleur", "orientation",
                "lock_x", "lock_y", "rayures", "couleur_rayures", "couleur_texte"]}
        for r in rects:
            for key, val in raw.items():
                if val == "" or val is None: continue
                r[key] = val
        draw_all(); sauvegarder()

    fig.canvas.get_tk_widget().winfo_toplevel().after(
        10, lambda: ouvrir_fenetre_rect(
            f"✏️ Editer groupe ({len(rects)} rectangles)", defaults, ok, groupe=True))

def editer_rectangle(rid):
    r = rect_by_id(rid)
    if r is None: return
    y_affich = NB_Y - r["y"] - r["largeur"]
    defaults = {"nom": r["nom"],
                "longueur": round(r["longueur"] * CASE_TO_CM_X, 3),
                "largeur":  round(r["largeur"]  * CASE_TO_CM_Y, 3),
                "x": round(r["x"] * CASE_TO_CM_X, 3),
                "y": round(y_affich * CASE_TO_CM_Y, 3),
                "couleur": r["couleur"], "orientation": r["orientation"],
                "lock_x": r.get("lock_x", False), "lock_y": r.get("lock_y", False),
                "rayures": r.get("rayures", False),
                "couleur_rayures": r.get("couleur_rayures", "#FFFFFF"),
                "couleur_texte": r.get("couleur_texte", "white")}
    def ok(result): r.update(result); draw_all(); sauvegarder()
    fig.canvas.get_tk_widget().winfo_toplevel().after(
        10, lambda: ouvrir_fenetre_rect(f"✏️ Éditer « {r['nom']} »", defaults, ok))

# ═══════════════════════════════════════════════════════════════════════════════
#  BOUTONS
# ═══════════════════════════════════════════════════════════════════════════════
def basculer_produit(prod, *args):
    if prod in state["produits_visibles"]:
        state["produits_visibles"].discard(prod)
    else:
        state["produits_visibles"].add(prod)
    btn = btn_filtres.get(prod)
    if btn:
        if prod in state["produits_visibles"]:
            btn.ax.set_facecolor(PRODUIT_COULEURS.get(prod, "#555555"))
        else:
            btn.ax.set_facecolor("#333333")
    draw_all()

def basculer_mode_connexion(*args):
    state["connecting"] = not state["connecting"]; state["conn_first"] = None
    if state["connecting"]:
        btn_conn.label.set_text("[ON] Connecter"); btn_conn.ax.set_facecolor("#1b5e20")
        ax.set_title(f"Carte {LONGUEUR_M}m x {LARGEUR_M}m  |  MODE CONNEXION : clic sur bord d'un rect",
                     fontsize=9, color="#FFD700")
    else:
        btn_conn.label.set_text("--- Connecter"); btn_conn.ax.set_facecolor("#0d47a1")
    fig.canvas.draw_idle()

def supprimer_selection(*args):
    global connexions, rectangles
    if state["conn_selected"] is not None:
        cid = state["conn_selected"]
        connexions = [c for c in connexions if c["id"] != cid]
        state["conn_selected"] = None
        draw_all(); sauvegarder(); return
    to_delete = set(state["group"])
    if state["selected"] is not None: to_delete.add(state["selected"])
    if to_delete:
        rectangles = [r for r in rectangles if r["id"] not in to_delete]
        connexions = [c for c in connexions
                      if c["r1"] not in to_delete and c["r2"] not in to_delete]
        state["selected"] = None; state["group"] = set()
        draw_all(); sauvegarder()

def dupliquer_rectangle(*args):
    import copy
    to_dup = set(state["group"])
    if state["selected"] is not None: to_dup.add(state["selected"])
    if not to_dup: return
    new_group = set()
    for rid in to_dup:
        r = rect_by_id(rid)
        if r is None: continue
        new_id = max((rec["id"] for rec in rectangles), default=-1) + 1
        new_r = copy.copy(r); new_r["id"] = new_id; new_r["nom"] = r["nom"] + " (copie)"
        new_r["x"] = r["x"] if r.get("lock_x", False) else min(r["x"] + 10, NB_X - r["longueur"])
        new_r["y"] = r["y"] if r.get("lock_y", False) else min(r["y"] + 10, NB_Y - r["largeur"])
        rectangles.append(new_r); new_group.add(new_id)
    state["selected"] = None; state["group"] = new_group
    draw_all(); sauvegarder()

# ═══════════════════════════════════════════════════════════════════════════════
#  CLAVIER
# ═══════════════════════════════════════════════════════════════════════════════
def on_key(event):
    if event.key in ("delete", "suppr", "backspace"):
        supprimer_selection()
    elif event.key in ("ctrl+d",):
        dupliquer_rectangle()

# ═══════════════════════════════════════════════════════════════════════════════
#  HIT-TESTS  (tolérances dynamiques en pixels via _tol_data)
# ═══════════════════════════════════════════════════════════════════════════════
def get_rect_at(mx, my):
    for r in reversed(rectangles):
        dx, dy = mat_to_display(r["x"], r["y"], r["largeur"])
        if dx <= mx <= dx + r["longueur"] and dy <= my <= dy + r["largeur"]:
            return r["id"]
    return None

def get_handle_at(mx, my):
    for hp in state["handles"]:
        x0, y0 = hp.get_xy(); w, h = hp.get_width(), hp.get_height()
        if x0 <= mx <= x0 + w and y0 <= my <= y0 + h: return hp.handle_name
    return None

def get_anchor_at(mx, my):
    """Détecte un nœud d'ancre. Tolérance = 10 px convertis en unités-données."""
    tol = _tol_data(10)
    for conn in connexions:
        if rect_by_id(conn["r1"]) is None or rect_by_id(conn["r2"]) is None:
            continue
        cid = conn["id"]
        p1 = anchor_point(cid, conn["r1"], conn["side1"])
        p2 = anchor_point(cid, conn["r2"], conn["side2"])
        if math.sqrt((mx - p1[0]) ** 2 + (my - p1[1]) ** 2) < tol: return (cid, "r1")
        if math.sqrt((mx - p2[0]) ** 2 + (my - p2[1]) ** 2) < tol: return (cid, "r2")
    return None

def get_segment_at(mx, my):
    """
    Détecte un segment de coude cliquable.
    - Tolérance = 8 px en unités-données.
    - Détection sur TOUTE la longueur du segment (pas seulement son milieu),
      ce qui rend le drag beaucoup plus facile à déclencher.
    """
    tol = _tol_data(8)
    EPS = 1e-6
    for cid, pts in state["coude_pts_cache"].items():
        seg_range = range(0, len(pts) - 1) if len(pts) <= 3 else range(1, len(pts) - 2)
        for i in seg_range:
            ax_, ay_ = pts[i]; bx, by = pts[i + 1]
            ddx = bx - ax_; ddy = by - ay_
            denom = ddx * ddx + ddy * ddy
            if denom < EPS: continue
            t = max(0.0, min(1.0, ((mx - ax_) * ddx + (my - ay_) * ddy) / denom))
            dist = math.sqrt((mx - ax_ - t * ddx) ** 2 + (my - ay_ - t * ddy) ** 2)
            if dist < tol:
                return (cid, i)
    return None

def get_conn_at(mx, my, tol=None):
    """
    Détecte un trait de connexion.
    - Tolérance par défaut = 6 px convertis en unités-données.
    - Distance point-segment sur chaque tronçon du chemin.
    """
    if tol is None:
        tol = _tol_data(6)
    EPS = 1e-6
    for cid, items in conn_lines.items():
        for item in items:
            if not hasattr(item, "get_xdata"): continue
            xs, ys = item.get_xdata(), item.get_ydata()
            if len(xs) < 2: continue
            for i in range(len(xs) - 1):
                ax_, ay_ = xs[i], ys[i]; bx, by = xs[i + 1], ys[i + 1]
                ddx = bx - ax_; ddy = by - ay_
                denom = ddx * ddx + ddy * ddy
                if denom < EPS: continue
                t = max(0.0, min(1.0, ((mx - ax_) * ddx + (my - ay_) * ddy) / denom))
                dist = math.sqrt((mx - ax_ - t * ddx) ** 2 + (my - ay_ - t * ddy) ** 2)
                if dist < tol:
                    return cid
    return None

# ═══════════════════════════════════════════════════════════════════════════════
#  ÉVÉNEMENTS SOURIS
# ═══════════════════════════════════════════════════════════════════════════════
def on_press(event):
    if event.inaxes != ax or event.xdata is None: return
    mx, my = event.xdata, event.ydata
    _rebuild_rects_idx()

    if state["connecting"] and event.button == 1:
        rid = get_rect_at(mx, my)
        if rid is not None:
            side = closest_side(rect_by_id(rid), mx, my)
            if state["conn_first"] is None:
                state["conn_first"] = (rid, side); r = rect_by_id(rid)
                ax.set_title(f"CONNEXION : départ sur '{r['nom']}' ({side}) -- cliquer le 2e rect",
                             fontsize=9, color="#FFD700"); fig.canvas.draw_idle()
            else:
                r1id, side1 = state["conn_first"]; r2id, side2 = rid, side
                state["conn_first"] = None
                if r1id == r2id: state["connecting"] = False; basculer_mode_connexion(); return
                cid = max((c["id"] for c in connexions), default=-1) + 1
                connexions.append({"id": cid, "r1": r1id, "side1": side1,
                                   "r2": r2id, "side2": side2, "nombre": 1, "produit": ""})
                state["connecting"] = False; basculer_mode_connexion(); draw_all(); sauvegarder()
        return

    if event.button == 3:
        rid = get_rect_at(mx, my)
        if rid is not None:
            state["selected"] = rid; draw_all(); editer_rectangle(rid); return
        cid = get_conn_at(mx, my)
        if cid is not None:
            state["conn_selected"] = cid; draw_all()
            fig.canvas.get_tk_widget().winfo_toplevel().after(10, lambda c=cid: editer_connexion(c))
            return
        if len(state["group"]) > 0:
            editer_groupe()
        return

    if event.button == 1:
        ctrl = (event.key in ("control", "ctrl+control") or
                getattr(event, "key", "") in ("control", "ctrl"))

        seg = get_segment_at(mx, my)
        if seg is not None:
            cid, idx = seg
            pts_now = state["coude_pts_cache"].get(cid, [])
            state["coude_dragging"] = True; state["coude_conn_id"] = cid
            state["coude_idx"] = idx; state["seg_pts_origin"] = list(pts_now)
            state["conn_selected"] = cid; return

        anchor = get_anchor_at(mx, my)
        if anchor is not None:
            cid, end = anchor; state["anchor_dragging"] = True
            state["anchor_conn_id"] = cid; state["anchor_end"] = end
            state["conn_selected"] = cid; draw_all(); return

        cid = get_conn_at(mx, my)
        if cid is not None:
            state["conn_selected"] = cid if state["conn_selected"] != cid else None
            draw_all(); return

        handle = get_handle_at(mx, my)
        if handle and state["selected"] is not None:
            state["resizing"] = True; state["resize_handle"] = handle
            r = rect_by_id(state["selected"])
            state["resize_origin"] = (mx, my, r["longueur"], r["largeur"]); return

        rid = get_rect_at(mx, my)
        if rid is not None:
            state["conn_selected"] = None
            if ctrl:
                if rid in state["group"]: state["group"].discard(rid)
                else: state["group"].add(rid)
                state["selected"] = rid
            else:
                if rid not in state["group"]:
                    state["group"] = set()
                state["selected"] = rid

            r = rect_by_id(rid)
            rdx, rdy = mat_to_display(r["x"], r["y"], r["largeur"])
            state["drag_offset"] = (mx - rdx, my - rdy)

            state["group_drag_offsets"] = {
                gid: (rect_by_id(gid)["x"] - r["x"],
                      rect_by_id(gid)["y"] - r["y"])
                for gid in state["group"] if rect_by_id(gid)
            }
            state["dragging"] = True
            draw_all()
        else:
            if ctrl:
                pass
            else:
                state["selected"] = None; state["conn_selected"] = None
                state["group"] = set()
                state["lasso"] = True; state["lasso_start"] = (mx, my)
                state["lasso_rect"] = None
            draw_all()

def on_release(event):
    if event.button != 1: return

    if state["coude_dragging"]:
        cid_done = state["coude_conn_id"]
        conn_done = next((c for c in connexions if c["id"] == cid_done), None)
        if conn_done: conn_done.pop("pts_cache", None)
        state["coude_dragging"] = False; state["coude_conn_id"] = None; state["coude_idx"] = None
        draw_all(); sauvegarder(); return

    if state["anchor_dragging"]:
        rid = get_rect_at(event.xdata, event.ydata) if event.xdata else None
        cid = state["anchor_conn_id"]; end = state["anchor_end"]
        conn = next((c for c in connexions if c["id"] == cid), None)
        if conn and rid is not None:
            other = conn["r2"] if end == "r1" else conn["r1"]
            if rid != other:
                side = closest_side(rect_by_id(rid), event.xdata, event.ydata)
                if end == "r1": conn["r1"] = rid; conn["side1"] = side
                else: conn["r2"] = rid; conn["side2"] = side
        if conn:
            conn.pop("waypoints_override", None); conn.pop("pts_cache", None)
        state["anchor_dragging"] = False; state["anchor_conn_id"] = None; state["anchor_end"] = None
        draw_all(); sauvegarder(); return

    if state["lasso"]:
        if state["lasso_rect"] is not None:
            try: state["lasso_rect"].remove()
            except: pass
            state["lasso_rect"] = None
        if state["lasso_start"] and event.xdata is not None:
            x0, y0 = state["lasso_start"]; x1, y1 = event.xdata, event.ydata
            lx0, lx1 = min(x0, x1), max(x0, x1); ly0, ly1 = min(y0, y1), max(y0, y1)
            for r in rectangles:
                rx0, ry0, rx1, ry1 = rect_display_bbox(r)
                cx = (rx0 + rx1) / 2; cy = (ry0 + ry1) / 2
                if lx0 <= cx <= lx1 and ly0 <= cy <= ly1:
                    state["group"].add(r["id"])
        state["lasso"] = False; state["lasso_start"] = None
        draw_all(); return

    if state["dragging"]:
        moved_rids = set(state["group_drag_offsets"].keys()) | (
            {state["selected"]} if state["selected"] is not None else set())
        for c in connexions:
            if c["r1"] in moved_rids or c["r2"] in moved_rids:
                c.pop("waypoints_override", None)
                c.pop("pts_cache", None)
        sauvegarder(avec_xlsx=False)
    elif state["resizing"]:
        sauvegarder(avec_xlsx=False)
    state["dragging"] = False; state["resizing"] = False; state["resize_origin"] = None

def on_motion(event):
    if event.inaxes != ax or event.xdata is None: return
    mx, my = event.xdata, event.ydata

    if state["dragging"] and state["selected"] is not None:
        r = rect_by_id(state["selected"])
        offx, offy = state["drag_offset"]
        target_dx = mx - offx
        target_dy = my - offy
        new_x_f, new_y_f = display_to_mat(target_dx, target_dy, r["largeur"])
        new_x = int(round(max(0, min(NB_X - r["longueur"], new_x_f))))
        new_y = int(round(max(0, min(NB_Y - r["largeur"],  new_y_f))))
        dx_move = 0 if r.get("lock_x", False) else new_x - r["x"]
        dy_move = 0 if r.get("lock_y", False) else new_y - r["y"]
        if dx_move == 0 and dy_move == 0:
            return
        r["x"] += dx_move; r["y"] += dy_move
        for gid in state["group_drag_offsets"]:
            if gid == state["selected"]: continue
            gr = rect_by_id(gid)
            if gr is None: continue
            if not gr.get("lock_x", False):
                gr["x"] = int(max(0, min(NB_X - gr["longueur"], gr["x"] + dx_move)))
            if not gr.get("lock_y", False):
                gr["y"] = int(max(0, min(NB_Y - gr["largeur"],  gr["y"] + dy_move)))
        moved_rids = set(state["group_drag_offsets"].keys()) | {state["selected"]}
        for c in connexions:
            if c["r1"] in moved_rids or c["r2"] in moved_rids:
                c.pop("pts_cache", None)
                if c.get("produit", "") in state["produits_visibles"]:
                    c.pop("waypoints_override", None)
        _partial_redraw_moved(moved_rids)
        return

    if state["coude_dragging"] and state["coude_conn_id"] is not None:
        cid    = state["coude_conn_id"]
        idx    = state["coude_idx"]
        origin = state["seg_pts_origin"]
        conn   = next((c for c in connexions if c["id"] == cid), None)
        if conn and origin and 0 <= idx < len(origin) - 1:
            EPS = 1e-6
            a = origin[idx]; b = origin[idx + 1]
            seg_is_h = abs(a[1] - b[1]) < EPS
            if seg_is_h:
                new_a = (a[0], my); new_b = (b[0], my)
            else:
                new_a = (mx, a[1]); new_b = (mx, b[1])

            def join_ortho(a, b):
                if abs(a[0] - b[0]) < EPS or abs(a[1] - b[1]) < EPS: return [b]
                return [(b[0], a[1]), b] if seg_is_h else [(a[0], b[1]), b]

            new_pts = list(origin[:idx])
            before = list(origin[:idx])
            after  = list(origin[idx + 2:])
            if before:
                for pt in join_ortho(before[-1], new_a): new_pts.append(pt)
            else:
                new_pts.append(new_a)
            if abs(new_a[0] - new_b[0]) > EPS or abs(new_a[1] - new_b[1]) > EPS:
                new_pts.append(new_b)
            if after:
                for pt in join_ortho(new_b, after[0]): new_pts.append(pt)
                new_pts.extend(after[1:])

            deduped = [new_pts[0]]
            for pt in new_pts[1:]:
                if abs(pt[0] - deduped[-1][0]) > EPS or abs(pt[1] - deduped[-1][1]) > EPS:
                    deduped.append(pt)
            conn["waypoints_override"] = deduped
            state["coude_pts_cache"][cid] = deduped
            saved_cache = dict(state["coude_pts_cache"]); saved_idx = state["coude_idx"]
            draw_all()
            state["coude_pts_cache"].update(saved_cache); state["coude_idx"] = saved_idx
        return

    if state["lasso"] and state["lasso_start"] is not None:
        x0, y0 = state["lasso_start"]
        if state["lasso_rect"] is not None:
            try: state["lasso_rect"].remove()
            except: pass
        lx = min(x0, mx); ly = min(y0, my); lw = abs(mx - x0); lh = abs(my - y0)
        lr = Rectangle((lx, ly), lw, lh, linewidth=1.5, edgecolor="#00FFFF",
                        facecolor="#00FFFF", alpha=0.08, linestyle="--", zorder=20)
        ax.add_patch(lr); state["lasso_rect"] = lr
        fig.canvas.draw_idle(); return

    if state["resizing"] and state["selected"] is not None:
        r = rect_by_id(state["selected"])
        ox, oy, ol, ow = state["resize_origin"]; h = state["resize_handle"]
        if h in ("right", "corner"):
            r["longueur"] = min(max(4, int(ol + (mx - ox))), NB_X - r["x"])
        if h in ("bottom", "corner"):
            r["largeur"]  = min(max(4, int(ow - (my - oy))), NB_Y - r["y"])
        draw_all(); return

    if state["anchor_dragging"] and state["anchor_conn_id"] is not None:
        clear_overlay(); cid = state["anchor_conn_id"]; end = state["anchor_end"]
        conn = next((c for c in connexions if c["id"] == cid), None)
        if conn:
            fixed_rid  = conn["r2"] if end == "r1" else conn["r1"]
            fixed_side = conn["side2"] if end == "r1" else conn["side1"]
            fixed_r = rect_by_id(fixed_rid)
            if fixed_r:
                fp = anchor_point(cid, fixed_rid, fixed_side)
                ln, = ax.plot([fp[0], mx], [fp[1], my], color="#AAAAAA",
                               linewidth=1.5, linestyle="--", zorder=9)
                dot, = ax.plot(mx, my, "o", color="#FFFFFF", markersize=5, zorder=10)
                state["overlay"].extend([ln, dot])
                rid_hover = get_rect_at(mx, my)
                if rid_hover is not None and rid_hover != fixed_rid:
                    r_h = rect_by_id(rid_hover)
                    dx_h, dy_h = mat_to_display(r_h["x"], r_h["y"], r_h["largeur"])
                    hl = Rectangle((dx_h, dy_h), r_h["longueur"], r_h["largeur"],
                                   linewidth=2, edgecolor="#00FF88", facecolor="none", zorder=9)
                    ax.add_patch(hl); state["overlay"].append(hl)
        fig.canvas.draw_idle()

# ═══════════════════════════════════════════════════════════════════════════════
#  FIGURE
# ═══════════════════════════════════════════════════════════════════════════════
def fmt_x(val, pos):
    v = val * CASE_TO_CM_X; return f"{v:.0f}" if v == int(v) else f"{v:.1f}"
def fmt_y(val, pos):
    v = val * CASE_TO_CM_Y; return f"{v:.0f}" if v == int(v) else f"{v:.1f}"

_L, _R, _B, _T = 0.07, 0.99, 0.14, 0.97
_plot_w_frac = _R - _L; _plot_h_frac = _T - _B
_fig_h = 8.0
_fig_w = _fig_h * (_plot_h_frac / _plot_w_frac) * (NB_X / NB_Y)
fig, ax = plt.subplots(figsize=(_fig_w, _fig_h))
plt.subplots_adjust(left=_L, right=_R, bottom=_B, top=_T)
ax.set_xlim(0, NB_X); ax.set_ylim(0, NB_Y)
ax.set_facecolor("#1a1a2e"); ax.tick_params(colors="white")
ax.set_xlabel("X (cm plan)", color="white"); ax.set_ylabel("Y (cm plan)", color="white")
for sp in ax.spines.values(): sp.set_edgecolor("#444")
fig.patch.set_facecolor("#12121e")
ax.set_xticks(np.arange(0, NB_X + 1, 40))
ax.set_yticks(np.arange(0, NB_Y + 1, 10))
ax.xaxis.set_major_formatter(mticker.FuncFormatter(fmt_x))
ax.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_y))
ax.grid(True, color="gray", linewidth=0.3, alpha=0.4)

def make_btn(pos, label, color, hover, cb):
    a = plt.axes(pos); b = Button(a, label, color=color, hovercolor=hover)
    b.label.set_color("white"); b.on_clicked(cb); return b

btn_add  = make_btn([0.01, 0.02, 0.13, 0.06], "+ Ajouter rect",  "#1b5e20", "#2e7d32", ajouter_rectangle_interactif)
btn_save = make_btn([0.16, 0.02, 0.13, 0.06], "  Sauvegarder",   "#0d47a1", "#1565c0", lambda e: sauvegarder())
btn_conn = make_btn([0.31, 0.02, 0.15, 0.06], "--- Connecter",   "#0d47a1", "#1565c0", basculer_mode_connexion)
btn_dup  = make_btn([0.48, 0.02, 0.14, 0.06], "Dupliquer",       "#004d40", "#00695c", dupliquer_rectangle)

btn_filtres = {}
filtre_labels    = {"A": "Prod A", "B": "Prod B", "C": "Prod C", "D": "Prod D"}
filtre_positions = [0.64, 0.71, 0.78, 0.85]
for prod, xpos in zip(["A", "B", "C", "D"], filtre_positions):
    col = PRODUIT_COULEURS.get(prod, "#555555")
    b = make_btn([xpos, 0.02, 0.06, 0.06], filtre_labels[prod], col, col,
                 lambda e, p=prod: basculer_produit(p))
    btn_filtres[prod] = b

ax.set_title(
    f"Carte {LONGUEUR_M}m x {LARGEUR_M}m  |  Clic gauche=select/drag  |  "
    f"Clic droit=éditer  |  Suppr=supprimer  |  Ctrl+D=dupliquer",
    fontsize=8, color="white")

fig.canvas.mpl_connect("button_press_event",   on_press)
fig.canvas.mpl_connect("button_release_event", on_release)
fig.canvas.mpl_connect("motion_notify_event",  on_motion)
fig.canvas.mpl_connect("key_press_event",      on_key)

def on_resize(event):
    fw = event.width / fig.dpi; fh = event.height / fig.dpi
    if fw <= 0 or fh <= 0: return
    plot_h = (_T - _B) * fh
    needed_plot_w = plot_h * (NB_X / NB_Y)
    new_right = _L + needed_plot_w / fw
    if 0.1 < new_right <= 1.0:
        plt.subplots_adjust(left=_L, right=new_right, bottom=_B, top=_T)
        fig.canvas.draw_idle()

fig.canvas.mpl_connect("resize_event", on_resize)

# ═══════════════════════════════════════════════════════════════════════════════
#  LANCEMENT
# ═══════════════════════════════════════════════════════════════════════════════
charger()
if not rectangles:
    rectangles += [
        {"id": 0, "nom": "Zone A", "longueur": 60, "largeur": 40, "x": 10, "y": 10,
         "couleur": "#4FC3F7", "orientation": "horizontal",
         "lock_x": False, "lock_y": False, "rayures": False,
         "couleur_rayures": "#FFFFFF", "couleur_texte": "white"},
        {"id": 1, "nom": "Zone B", "longueur": 40, "largeur": 60, "x": 200, "y": 50,
         "couleur": "#81C784", "orientation": "vertical",
         "lock_x": False, "lock_y": False, "rayures": False,
         "couleur_rayures": "#FFFFFF", "couleur_texte": "white"},
    ]
draw_all()
plt.show()